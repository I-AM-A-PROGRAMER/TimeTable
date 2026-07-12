"""
Parses the KIIT timetable Excel file into a SQLite database.

Usage:
    python build_db.py [path/to/excel.xlsx]

If no path is given, it auto-detects the first .xlsx in this folder.
"""
import glob
import os
import re
import sqlite3
import sys

import openpyxl

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "timetable.db")

# Period definitions extracted from the Excel header.
# Each period is a 1-hour block; P1 starts at 08:00.
PERIODS = [
    ("P1", "08:00", "09:00"),
    ("P2", "09:00", "10:00"),
    ("P3", "10:00", "11:00"),
    ("P4", "11:00", "12:00"),
    ("P5", "12:00", "13:00"),
    ("P6", "13:00", "14:00"),
    ("P7", "14:00", "15:00"),
    ("P8", "15:00", "16:00"),
    ("P9", "16:00", "17:00"),
    ("P10", "17:00", "18:00"),
]

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def find_excel(arg):
    if arg:
        return arg
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = sorted(glob.glob(os.path.join(here, "*.xlsx")))
    candidates = [c for c in candidates if not os.path.basename(c).startswith("~$")]
    if not candidates:
        sys.exit("No .xlsx file found. Place the timetable in this folder or pass its path.")
    return candidates[0]


def create_schema(conn):
    conn.executescript(
        """
        DROP TABLE IF EXISTS classes;
        DROP TABLE IF EXISTS sections;
        DROP TABLE IF EXISTS periods;

        CREATE TABLE periods (
            id   INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            start_time TEXT NOT NULL,
            end_time   TEXT NOT NULL
        );

        CREATE TABLE sections (
            id           INTEGER PRIMARY KEY,
            code         TEXT NOT NULL UNIQUE,
            department   TEXT,
            semester     TEXT,
            course_groups TEXT
        );

        CREATE TABLE classes (
            id         INTEGER PRIMARY KEY,
            section_id INTEGER NOT NULL,
            day        TEXT NOT NULL,
            period_id  INTEGER NOT NULL,
            subject    TEXT,
            teacher    TEXT,
            room       TEXT,
            FOREIGN KEY (section_id) REFERENCES sections(id),
            FOREIGN KEY (period_id)  REFERENCES periods(id)
        );

        CREATE INDEX idx_classes_section ON classes(section_id);
        CREATE INDEX idx_classes_day ON classes(section_id, day);
        """
    )


def seed_periods(conn):
    for i, (name, start, end) in enumerate(PERIODS, start=1):
        conn.execute(
            "INSERT INTO periods (id, name, start_time, end_time) VALUES (?,?,?,?)",
            (i, name, start, end),
        )


def split_section_banner(value):
    """Banner cell looks like: 'Sem 5 | CS-S5 | CS1'. Split into parts.

    Only accepts banners for Semester 5; returns None for other semesters.
    """
    if not value:
        return None
    parts = [p.strip() for p in str(value).split("|")]
    # Expect at least 3 parts: semester, department, code
    if len(parts) >= 3 and re.search(r"Sem\s*5", parts[0]):
        return parts[0], parts[1], parts[2]
    return None


def parse_cell(value):
    """A class cell is 3 lines: subject, teacher, room. Returns (subject, teacher, room) or None."""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    if not lines:
        return None
    subject = lines[0] if len(lines) > 0 else None
    teacher = lines[1] if len(lines) > 1 else None
    room = lines[2] if len(lines) > 2 else None
    return subject, teacher, room


def is_valid_day(value):
    if not value:
        return False
    return str(value).strip().title() in DAYS


def parse_workbook(conn, path):
    wb = openpyxl.load_workbook(path, data_only=True)
    # Use the grid sheet (first one).
    ws = wb.worksheets[0]
    max_row = ws.max_row
    max_col = min(ws.max_column, 12)  # cols A..L = 1..12

    current_section_id = None

    for r in range(2, max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value

        # Banner row? (column A contains "Sem ... | DEPT | CODE")
        if a and re.search(r"Sem\s*\d", str(a)):
            banner = split_section_banner(str(a))
            if banner:
                semester, department, code = banner
                course_groups = str(b) if b else None
                cur = conn.execute(
                    "INSERT OR IGNORE INTO sections (code, department, semester, course_groups) VALUES (?,?,?,?)",
                    (code, department, semester, course_groups),
                )
                if cur.rowcount == 0:
                    # already exists; fetch id
                    row = conn.execute(
                        "SELECT id FROM sections WHERE code=?", (code,)
                    ).fetchone()
                    current_section_id = row[0] if row else None
                else:
                    current_section_id = cur.lastrowid
            continue

        # Data row for a section/day?
        if current_section_id and is_valid_day(b):
            day = str(b).strip().title()
            # Cols 3..12 = P1..P10
            for c in range(3, 13):
                period_idx = c - 3  # 0..9
                if period_idx >= len(PERIODS):
                    break
                cell = parse_cell(ws.cell(r, c).value)
                if cell:
                    subject, teacher, room = cell
                    period_id = period_idx + 1
                    conn.execute(
                        "INSERT INTO classes (section_id, day, period_id, subject, teacher, room) VALUES (?,?,?,?,?,?)",
                        (current_section_id, day, period_id, subject, teacher, room),
                    )


def main():
    path = find_excel(sys.argv[1] if len(sys.argv) > 1 else None)
    print(f"Parsing: {path}")
    conn = sqlite3.connect(DB_PATH)
    try:
        create_schema(conn)
        seed_periods(conn)
        parse_workbook(conn, path)
        conn.commit()

        sec_count = conn.execute("SELECT COUNT(*) FROM sections").fetchone()[0]
        cls_count = conn.execute("SELECT COUNT(*) FROM classes").fetchone()[0]
        print(f"Done. {sec_count} sections, {cls_count} classes imported.")
        print(f"Database: {DB_PATH}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
